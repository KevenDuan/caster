import sys
import time
import os
import threading
import numpy as np  
import pyqtgraph as pg 
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QGridLayout, QLabel, QLineEdit, 
                             QPushButton, QGroupBox, QMessageBox, QStatusBar, 
                             QSizePolicy, QTabWidget, QFileDialog, QSpinBox)
from PyQt5.QtCore import QTimer, QThread, pyqtSignal, Qt, QSettings
from PyQt5.QtGui import QPixmap, QFont
from pymodbus.client import ModbusTcpClient
from pymodbus.payload import BinaryPayloadDecoder, BinaryPayloadBuilder
from pymodbus.constants import Endian

# --- Excel 模块 ---
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import logging
logging.getLogger("pymodbus").setLevel(logging.ERROR)

# ==========================================
# 1. PLC 轮询子线程
# ==========================================
class PlcPollerThread(QThread):
    data_updated = pyqtSignal(dict)
    connection_status = pyqtSignal(bool)

    def __init__(self, ip='192.168.6.6', port=502):
        super().__init__()
        self.ip = ip
        self.port = port
        self.is_running = True
        self.client = ModbusTcpClient(self.ip, port=self.port)

    def run(self):
        while self.is_running:
            if not self.client.connect():
                self.connection_status.emit(False)
                time.sleep(1) 
                continue
            
            self.connection_status.emit(True)
            try:
                # 读取 D0 ~ D46，共 47 个寄存器
                result = self.client.read_holding_registers(address=0, count=47)
                if not result.isError():
                    decoder = BinaryPayloadDecoder.fromRegisters(
                        result.registers, byteorder=Endian.BIG, wordorder=Endian.LITTLE
                    )
                    d0_status = decoder.decode_16bit_uint()       
                    decoder.skip_bytes(2)                         
                    d2_speed = decoder.decode_32bit_float()      
                    d4_cycle = decoder.decode_16bit_uint()        
                    decoder.skip_bytes(2)                         
                    d6_time = decoder.decode_32bit_float()        
                    d8_mileage = decoder.decode_32bit_float()
                    
                    d10_temp1 = decoder.decode_16bit_int()        
                    decoder.skip_bytes(50)                        
                    d36_temp2 = decoder.decode_16bit_int()        
                    
                    # 【核心修改点：重命名为轮旋转次数】
                    decoder.skip_bytes(2)                         
                    d38_wheel_revolutions = decoder.decode_32bit_int() 
                    
                    d40_weight1 = decoder.decode_16bit_int()      
                    decoder.skip_bytes(2)                         
                    d42_weight2 = decoder.decode_16bit_int()      

                    decoder.skip_bytes(6)                         
                    d46_control = decoder.decode_16bit_int()      

                    data = {
                        'status': d0_status,
                        'speed': round(d2_speed, 2),
                        'cycle': int(d4_cycle),  
                        'time': int(d6_time),    
                        'mileage': round(d8_mileage, 2),
                        'temp1': round(d10_temp1 / 10.0, 1),    
                        'temp2': round(d36_temp2 / 10.0, 1),
                        'weight1': d40_weight1, 
                        'weight2': d42_weight2,
                        'wheel_revolutions': d38_wheel_revolutions,     
                        'control_d46': d46_control  
                    }
                    self.data_updated.emit(data)
            except Exception as e:
                print(f"Read Exception: {e}")
                self.client.close()
            
            time.sleep(0.1)

    def stop(self):
        self.is_running = False
        self.client.close()

# ==========================================
# 2. 主窗口 UI
# ==========================================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.plc_ip = '192.168.6.6'
        self.current_plc_status = -1 
        
        self.time_history = []
        self.temp1_history = []
        self.temp2_history = []
        self.weight1_history = []
        self.weight2_history = []
        
        self.current_temp1 = 0 
        self.current_temp2 = 0 
        self.current_weight1 = 0 
        self.current_weight2 = 0 
        self.current_wheel_revolutions = 0  # 【核心修改点】
        
        self.plot_start_time = None
        self.test_start_time = ""
        self.test_end_time = ""
        self.ignore_sync_lock = False
        self.report_snapshot = None       
        self.last_plc_status = -1         

        self.settings = QSettings('KangruiConfig', 'WanliApp')

        self.init_ui()
        self.load_settings() 
        self.start_plc_thread()

    def create_plot(self, title, y_label, y_unit, color_hex):
        plot_widget = pg.PlotWidget(title=title)
        plot_widget.setBackground('#303133') 
        plot_widget.getAxis('left').enableAutoSIPrefix(False)
        plot_widget.setLabel('left', y_label, units=y_unit, color='#DCDFE6', size='12pt')
        plot_widget.setLabel('bottom', '运行时间', units='s', color='#DCDFE6', size='12pt')
        plot_widget.showGrid(x=True, y=True, alpha=0.3)
        font = QFont("Arial", 10)
        plot_widget.getAxis('bottom').setTickFont(font)
        plot_widget.getAxis('left').setTickFont(font)
        plot_widget.getAxis('bottom').setPen('#DCDFE6')
        plot_widget.getAxis('left').setPen('#DCDFE6')
        curve = plot_widget.plot(pen=pg.mkPen(color=color_hex, width=2)) 
        return plot_widget, curve

    def init_ui(self):
        self.setWindowTitle('康瑞控制测试系统 -- 万里脚轮')
        self.resize(1150, 720) 
        self.setStyleSheet("""
            QMainWindow { background-color: #F0F2F5; }
            QTabWidget::pane { border: 1px solid #DCDFE6; border-radius: 5px; background: white; }
            QTabBar::tab { background: #E4E7ED; padding: 10px 30px; margin-right: 2px; font-weight: bold; border-top-left-radius: 4px; border-top-right-radius: 4px; }
            QTabBar::tab:selected { background: #409EFF; color: white; }
            QGroupBox { font-weight: bold; border: 1px solid #DCDFE6; border-radius: 5px; margin-top: 15px; padding-top: 15px;}
            QGroupBox::title { subcontrol-origin: margin; left: 10px; color: #409EFF;}
            QLabel { font-size: 14px; }
            QLineEdit { padding: 4px; border: 1px solid #DCDFE6; border-radius: 4px; background: white; min-height: 22px;}
            QSpinBox { padding: 4px; border: 1px solid #DCDFE6; border-radius: 4px; background: white; min-height: 22px; font-size: 13px;}
            QSpinBox::up-button, QSpinBox::down-button { width: 16px; }
            QPushButton { padding: 8px 15px; background-color: #409EFF; color: white; border: none; border-radius: 4px; font-weight: bold; font-size: 14px;}
            QPushButton:hover { background-color: #66B1FF; }
            QPushButton:disabled { background-color: #C0C4CC; color: #F0F2F5;}
            QPushButton#btn_danger { background-color: #F56C6C; }
            QPushButton#btn_danger:hover { background-color: #F89898; }
            QPushButton#btn_report { background-color: #E6A23C; }
            QPushButton#btn_report:hover { background-color: #EBB563; }
            QPushButton#btn_measure:checked { background-color: #67C23A; } 
            QPushButton#btn_measure:checked:hover { background-color: #85CE61; }
        """)

        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        self.tabs = QTabWidget()
        self.tab_main = QWidget()
        self.tab_graph = QWidget()
        self.tabs.addTab(self.tab_main, "🎛️ 参数界面")
        self.tabs.addTab(self.tab_graph, "📈 数据曲线")
        main_layout.addWidget(self.tabs)

        tab1_layout = QHBoxLayout(self.tab_main)
        tab1_layout.setSpacing(20)

        col1_layout = QVBoxLayout()
        col1_layout.addStretch() 
        self.lbl_logo = QLabel()
        self.lbl_logo.setAlignment(Qt.AlignCenter) 
        logo_path = 'jiaolun.png'
        if os.path.exists(logo_path):
            pixmap = QPixmap(logo_path)
            if not pixmap.isNull():
                self.lbl_logo.setPixmap(pixmap.scaledToWidth(260, Qt.SmoothTransformation))
            else:
                self.lbl_logo.setText("Logo 错误")
        else:
            self.lbl_logo.setText("未找到 Logo")
            self.lbl_logo.setStyleSheet("color: #909399; font-weight: bold;")
        col1_layout.addWidget(self.lbl_logo)
        col1_layout.addStretch()

        col2_layout = QVBoxLayout()
        col2_layout.setSpacing(10) 
        
        grp_info = QGroupBox("产品信息")
        layout_info = QGridLayout()
        layout_info.setSpacing(8)
        self.inputs_info = {}
        info_labels = ["产品名称", "型号", "编号", "轮径 (mm)", "材质", "硬度", "备注"]
        for i, label in enumerate(info_labels):
            layout_info.addWidget(QLabel(label + ":"), i, 0)
            le = QLineEdit()
            self.inputs_info[label] = le
            layout_info.addWidget(le, i, 1)
            # 其实这里不影响，因为链接的函数我不检查是否写了参数
            le.textChanged.connect(self.check_import_ready)
        grp_info.setLayout(layout_info)
        col2_layout.addWidget(grp_info)

        grp_params = QGroupBox("参数设置")
        layout_params = QGridLayout()
        layout_params.setSpacing(8)
        
        self.inputs_params = {}
        self.time_inputs = {} 
        
        self.param_dict = {
            "测试里程 (km)": "line", "测试速度 (km/h)": "line",
            "测试时长": "hms", "间隔时间": "hms",
            "障碍数量 (PCS)": "line", "障碍次数 (PCS)": "line",
            "障碍高度 (mm)": "line", "承载重量 (kg)": "line",
            "承载温度 (℃)": "line"
        }
        
        row = 0
        for name, p_type in self.param_dict.items():
            layout_params.addWidget(QLabel(name + ":"), row, 0)
            if p_type == "line":
                le = QLineEdit()
                self.inputs_params[name] = le
                layout_params.addWidget(le, row, 1)
                le.textChanged.connect(self.check_import_ready)
            elif p_type == "hms":
                hms_widget = QWidget()
                hms_layout = QHBoxLayout(hms_widget)
                hms_layout.setContentsMargins(0, 0, 0, 0)
                hms_layout.setSpacing(5)
                spin_h = QSpinBox()
                spin_h.setRange(0, 9999)
                spin_h.setSuffix(" 时")
                spin_m = QSpinBox()
                spin_m.setRange(0, 59)
                spin_m.setSuffix(" 分")
                spin_s = QSpinBox()
                spin_s.setRange(0, 59)
                spin_s.setSuffix(" 秒")
                for sp in [spin_h, spin_m, spin_s]:
                    sp.valueChanged.connect(self.check_import_ready)
                    hms_layout.addWidget(sp)
                layout_params.addWidget(hms_widget, row, 1)
                self.time_inputs[name] = (spin_h, spin_m, spin_s)
            row += 1
            
        grp_params.setLayout(layout_params)
        col2_layout.addWidget(grp_params)
        col2_layout.addStretch()

        col3_layout = QVBoxLayout()
        col3_layout.setSpacing(10)
        
        self.lbl_main_title = QLabel("万 里 脚 轮")
        self.lbl_main_title.setAlignment(Qt.AlignCenter)
        self.lbl_main_title.setStyleSheet("""
            font-size: 48px; 
            font-family: '华文行楷', 'STXingkai', '楷体', 'KaiTi', '隶书', 'LiSu', serif; 
            font-weight: bold; color: #2C3E50; margin-top: 5px; margin-bottom: 5px; letter-spacing: 6px;
        """)
        col3_layout.addWidget(self.lbl_main_title)
        
        grp_data = QGroupBox("实时数据监测")
        layout_data = QGridLayout()
        layout_data.setSpacing(8) 
        self.lbl_displays = {}
        
        data_configs = [
            ("speed", "速度 (km/h):"), ("cycle", "循环 (时:分:秒):"), 
            ("time", "时间 (时:分:秒):"), ("mileage", "里程 (km):"), 
            ("temp1", "工位1温度(℃):"), ("temp2", "工位2温度(℃):"),
            ("weight1", "工位1压力(kg):"), ("weight2", "工位2压力(kg):")
        ]
        
        lcd_font = QFont("Arial", 16, QFont.Bold)
        for i, (key, text) in enumerate(data_configs):
            r = i // 2
            c = (i % 2) * 2
            layout_data.addWidget(QLabel(text), r, c)
            lbl_val = QLabel("0.0" if key not in ['time', 'cycle'] else "00:00:00")
            lbl_val.setFont(lcd_font)
            lbl_val.setStyleSheet("color: #67C23A; background: #303133; padding: 4px; border-radius: 4px; min-width: 85px;")
            lbl_val.setAlignment(Qt.AlignCenter)
            self.lbl_displays[key] = lbl_val
            layout_data.addWidget(lbl_val, r, c + 1)
            
        grp_data.setLayout(layout_data)
        col3_layout.addWidget(grp_data) 

        grp_manual = QGroupBox("手动控制 (点动)")
        layout_manual = QGridLayout()
        layout_manual.setSpacing(10)
        layout_manual.setAlignment(Qt.AlignCenter) 
        
        btn_fwd = QPushButton("转盘正转")
        btn_fwd.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        btn_fwd.setMinimumHeight(40)
        btn_fwd.pressed.connect(lambda: self.write_manual_register(12, 1))
        btn_fwd.released.connect(lambda: self.write_manual_register(12, 0))
        
        btn_rev = QPushButton("转盘反转")
        btn_rev.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        btn_rev.setMinimumHeight(40)
        btn_rev.pressed.connect(lambda: self.write_manual_register(14, 1))
        btn_rev.released.connect(lambda: self.write_manual_register(14, 0))

        btn_weight1 = QPushButton("砝码1测量")
        btn_weight1.setObjectName("btn_measure")  # 赋予专属ID，配合样式表变色
        btn_weight1.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        btn_weight1.setMinimumHeight(40)
        btn_weight1.setCheckable(True)  # 开启自锁模式：点一下开启，再点一下关闭
        # checked 是一个布尔值(True/False)，True时写入1，False时写入0
        btn_weight1.toggled.connect(lambda checked: self.write_manual_register(48, 1 if checked else 0))

        btn_weight2 = QPushButton("砝码2测量")
        btn_weight2.setObjectName("btn_measure")  # 赋予专属ID，配合样式表变色
        btn_weight2.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        btn_weight2.setMinimumHeight(40)
        btn_weight2.setCheckable(True)  # 开启自锁模式
        btn_weight2.toggled.connect(lambda checked: self.write_manual_register(50, 1 if checked else 0))
        
        btn_clear = QPushButton("数据清除")
        btn_clear.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        btn_clear.setMinimumHeight(40)
        btn_clear.setObjectName("btn_danger")
        btn_clear.pressed.connect(lambda: self.write_manual_register(16, 1))
        btn_clear.released.connect(lambda: self.write_manual_register(16, 0))

        btn_report = QPushButton("测试报表预览") 
        btn_report.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        btn_report.setMinimumHeight(40)
        btn_report.setObjectName("btn_report")
        btn_report.clicked.connect(self.show_test_report_dialog)

        layout_manual.addWidget(btn_fwd, 0, 0)
        layout_manual.addWidget(btn_rev, 0, 1)
        layout_manual.addWidget(btn_weight1, 1, 0)
        layout_manual.addWidget(btn_weight2, 1, 1)
        layout_manual.addWidget(btn_clear, 2, 0)
        layout_manual.addWidget(btn_report, 2, 1)
        
        grp_manual.setLayout(layout_manual)
        col3_layout.addWidget(grp_manual) 

        self.btn_import = QPushButton("⬇️ 导入方案到 PLC")
        self.btn_import.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_import.setMinimumHeight(45) 
        self.btn_import.setEnabled(False)
        self.btn_import.clicked.connect(self.import_params)
        col3_layout.addWidget(self.btn_import)

        self.btn_print = QPushButton("🖨️ 生成报表")
        self.btn_print.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_print.setMinimumHeight(45) 
        self.btn_print.clicked.connect(self.export_to_excel)
        col3_layout.addWidget(self.btn_print)

        self.btn_start_stop = QPushButton("▶️ 启动")
        self.btn_start_stop.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_start_stop.setMinimumHeight(45) 
        self.btn_start_stop.setStyleSheet("""
            QPushButton {
                background-color: #67C23A; 
                color: white;
                border-radius: 4px; 
                font-weight: bold;
                font-size: 16px;
            }
            QPushButton:hover { background-color: #85CE61; }
        """)
        self.btn_start_stop.clicked.connect(self.toggle_start_stop)
        col3_layout.addWidget(self.btn_start_stop)

        self.lbl_tip = QLabel("技术支持：康瑞智能化科技有限公司")
        self.lbl_tip.setStyleSheet("color: #F56C6C; font-weight: bold; font-size: 13px; margin-top: 10px;")
        self.lbl_tip.setAlignment(Qt.AlignCenter)
        col3_layout.addWidget(self.lbl_tip)

        self.lbl_detailed_status = QLabel("设备状态: 正在等待连接...")
        self.lbl_detailed_status.setStyleSheet("font-size: 18px; font-weight: bold; margin-top: 5px;")
        self.lbl_detailed_status.setAlignment(Qt.AlignCenter)
        col3_layout.addWidget(self.lbl_detailed_status)

        col3_layout.addStretch()

        tab1_layout.addLayout(col1_layout, stretch=3) 
        tab1_layout.addLayout(col2_layout, stretch=3)     
        tab1_layout.addLayout(col3_layout, stretch=4)      

        # ==========================================
        # TAB 2: 图表布局 
        # ==========================================
        tab2_layout = QGridLayout(self.tab_graph)
        tab2_layout.setSpacing(15)
        self.plot_t1, self.curve_t1 = self.create_plot("工位 1 温度实时监测 (30s/点)", "温度", "℃", "#F56C6C")
        self.plot_t2, self.curve_t2 = self.create_plot("工位 2 温度实时监测 (30s/点)", "温度", "℃", "#E6A23C")
        self.plot_w1, self.curve_w1 = self.create_plot("工位 1 压力实时监测 (30s/点)", "重量", "kg", "#409EFF")
        self.plot_w2, self.curve_w2 = self.create_plot("工位 2 压力实时监测 (30s/点)", "重量", "kg", "#67C23A")
        tab2_layout.addWidget(self.plot_t1, 0, 0) 
        tab2_layout.addWidget(self.plot_t2, 0, 1) 
        tab2_layout.addWidget(self.plot_w1, 1, 0) 
        tab2_layout.addWidget(self.plot_w2, 1, 1) 

        # === 状态栏 ===
        self.statusBar = QStatusBar()
        self.setStatusBar(self.statusBar)
        self.lbl_time = QLabel()
        self.lbl_plc_status = QLabel("PLC 状态: 未连接")
        self.statusBar.addPermanentWidget(self.lbl_plc_status)
        self.statusBar.addPermanentWidget(self.lbl_time)

        self.timer_sys = QTimer(self)
        self.timer_sys.timeout.connect(self.update_local_time)
        self.timer_sys.start(1000)

        self.timer_plot = QTimer(self)
        self.timer_plot.timeout.connect(self.update_plot_canvas)
        self.timer_plot.start(30000) 

    # ================= 核心持久化存储 =================
    def load_settings(self):
        for key, le in self.inputs_info.items():
            val = self.settings.value(f"info/{key}", "")
            if val: le.setText(str(val))
            
        for key, p_type in self.param_dict.items():
            if p_type == "line":
                val = self.settings.value(f"param/{key}", "")
                if val: self.inputs_params[key].setText(str(val))
            elif p_type == "hms":
                h = self.settings.value(f"time/{key}_h", 0, type=int)
                m = self.settings.value(f"time/{key}_m", 0, type=int)
                s = self.settings.value(f"time/{key}_s", 0, type=int)
                self.time_inputs[key][0].setValue(h)
                self.time_inputs[key][1].setValue(m)
                self.time_inputs[key][2].setValue(s)

    def save_settings(self):
        for key, le in self.inputs_info.items():
            self.settings.setValue(f"info/{key}", le.text())
        for key, p_type in self.param_dict.items():
            if p_type == "line":
                self.settings.setValue(f"param/{key}", self.inputs_params[key].text())
            elif p_type == "hms":
                self.settings.setValue(f"time/{key}_h", self.time_inputs[key][0].value())
                self.settings.setValue(f"time/{key}_m", self.time_inputs[key][1].value())
                self.settings.setValue(f"time/{key}_s", self.time_inputs[key][2].value())

    def get_param_value_str(self, name):
        if name in self.inputs_params:
            return self.inputs_params[name].text()
        elif name in self.time_inputs:
            h, m, s = self.time_inputs[name]
            return f"{h.value()}时 {m.value()}分 {s.value()}秒"
        return ""

    def update_local_time(self):
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        self.lbl_time.setText(f"系统时间: {now}")

    def update_plot_canvas(self):
        if self.current_plc_status == -1: return 
        if self.plot_start_time is None:
            self.plot_start_time = time.time()
        
        elapsed_time = time.time() - self.plot_start_time
        
        MAX_POINTS = 2880
        if len(self.time_history) > MAX_POINTS:
            self.time_history.pop(0)
            self.temp1_history.pop(0)
            self.temp2_history.pop(0)
            self.weight1_history.pop(0)
            self.weight2_history.pop(0)
            
        self.time_history.append(elapsed_time)
        self.temp1_history.append(self.current_temp1)
        self.temp2_history.append(self.current_temp2)
        self.weight1_history.append(self.current_weight1)
        self.weight2_history.append(self.current_weight2)
        
        self.curve_t1.setData(np.array(self.time_history), np.array(self.temp1_history))
        self.curve_t2.setData(np.array(self.time_history), np.array(self.temp2_history))
        self.curve_w1.setData(np.array(self.time_history), np.array(self.weight1_history))
        self.curve_w2.setData(np.array(self.time_history), np.array(self.weight2_history))

    def check_import_ready(self):
        all_filled = all(le.text().strip() != "" for le in self.inputs_params.values()) 
        is_stopped = (self.current_plc_status == 0)
        self.btn_import.setEnabled(all_filled and is_stopped)
        if not is_stopped and self.current_plc_status != -1:
            self.btn_import.setText("导入方案 (需停止状态)")
            self.btn_import.setStyleSheet("") 
        else:
            self.btn_import.setText("⬇️ 导入方案到 PLC")
            self.btn_import.setStyleSheet("") 

    def capture_report_snapshot(self):
        self.report_snapshot = {
            'speed': self.lbl_displays['speed'].text(),
            'cycle': self.lbl_displays['cycle'].text(),
            'time': self.lbl_displays['time'].text(),
            'mileage': self.lbl_displays['mileage'].text(),
            'temp1': self.lbl_displays['temp1'].text(),
            'temp2': self.lbl_displays['temp2'].text(),
            'weight1': self.lbl_displays['weight1'].text(),
            'weight2': self.lbl_displays['weight2'].text(),
            'wheel_revolutions': self.current_wheel_revolutions
        }

    def update_plc_ui(self, data):
        self.current_temp1 = data.get('temp1', 0)
        self.current_temp2 = data.get('temp2', 0)
        self.current_weight1 = data.get('weight1', 0)
        self.current_weight2 = data.get('weight2', 0)
        
        self.current_wheel_revolutions = data.get('wheel_revolutions', 0)

        for key in ['speed', 'mileage', 'temp1', 'temp2', 'weight1', 'weight2']:
            new_text = str(data.get(key, 0))
            if self.lbl_displays[key].text() != new_text:
                self.lbl_displays[key].setText(new_text)
                
        for key in ['time', 'cycle']:
            total_seconds = int(data.get(key, 0))
            h = total_seconds // 3600
            m = (total_seconds % 3600) // 60
            s = total_seconds % 60
            new_text = f"{h:02d}:{m:02d}:{s:02d}"
            if self.lbl_displays[key].text() != new_text:
                self.lbl_displays[key].setText(new_text)
        
        new_status = data['status']
        if self.last_plc_status == 1 and new_status != 1:
            if self.report_snapshot is None:
                self.capture_report_snapshot()
                
        self.last_plc_status = new_status
        self.current_plc_status = new_status
        
        status_map = {
            0: ("已停止", "#909399"), 1: ("运行中", "#67C23A"), 2: ("暂停中", "#E6A23C"), 
            4: ("里程已到达", "#409EFF"), 5: ("1号温度过高", "#F56C6C"), 6: ("2号温度过高", "#F56C6C"),
            7: ("1号砝码异常", "#F56C6C"), 8: ("2号砝码异常", "#F56C6C"),
            9: ("障碍次数已到达", "#409EFF"), 10: ("2号升降电机异常", "#F56C6C"),
            11: ("1号升降电机异常", "#F56C6C"),
        }
        text, color = status_map.get(self.current_plc_status, (f"未知状态代码:{self.current_plc_status}", "#909399"))
        self.lbl_plc_status.setText(f"PLC 状态: <span style='color:{color}; font-weight:bold;'>{text}</span>")
        self.lbl_detailed_status.setText(f"设备状态: <span style='color:{color};'>{text}</span>")
        self.check_import_ready()

        if not self.ignore_sync_lock:
            control_val = data.get('control_d46', 0)
            if control_val == 1 and self.btn_start_stop.text() == "▶️ 启动":
                self.update_btn_style(is_running=True)
            elif control_val == 0 and self.btn_start_stop.text() == "⏹️ 停止":
                self.update_btn_style(is_running=False)

    def update_btn_style(self, is_running):
        if is_running:
            self.btn_start_stop.setText("⏹️ 停止")
            self.btn_start_stop.setStyleSheet("QPushButton { background-color: #F56C6C; color: white; border-radius: 4px; font-weight: bold; font-size: 16px; } QPushButton:hover { background-color: #F89898; }")
        else:
            self.btn_start_stop.setText("▶️ 启动")
            self.btn_start_stop.setStyleSheet("QPushButton { background-color: #67C23A; color: white; border-radius: 4px; font-weight: bold; font-size: 16px; } QPushButton:hover { background-color: #85CE61; }")

    def toggle_start_stop(self):
        self.ignore_sync_lock = True
        QTimer.singleShot(1500, self.release_sync_lock)
        if self.btn_start_stop.text() == "▶️ 启动":
            self.report_snapshot = None  
            self.write_manual_register(46, 1)  
            self.update_btn_style(is_running=True) 
        else:
            self.capture_report_snapshot() 
            self.write_manual_register(46, 0)  
            self.update_btn_style(is_running=False) 

    def release_sync_lock(self):
        self.ignore_sync_lock = False

    def show_test_report_dialog(self):
        if self.report_snapshot is None:
            self.capture_report_snapshot()
        snap = self.report_snapshot

        set_obs_str = self.get_param_value_str("障碍数量 (PCS)")
        try:
            set_obs_val = int(set_obs_str) if set_obs_str.strip() else 0
        except ValueError:
            set_obs_val = 0
            
        # 核心修改：实测总次数 = D38读取的轮旋转次数 * 障碍数量
        actual_total_obstacles = snap['wheel_revolutions'] * set_obs_val
        display_actual_obs = f"{actual_total_obstacles} PCS" if actual_total_obstacles > 0 else "无"

        th, tm, ts = self.time_inputs["测试时长"]
        run_s = th.value() * 3600 + tm.value() * 60 + ts.value()
        run_min = run_s / 60.0
        run_min_str = f"{int(run_min)}" if run_min.is_integer() else f"{run_min:.1f}"
        
        ih, im, i_s = self.time_inputs["间隔时间"]
        pause_s = ih.value() * 3600 + im.value() * 60 + i_s.value()
        pause_min = pause_s / 60.0
        pause_min_str = f"{int(pause_min)}" if pause_min.is_integer() else f"{pause_min:.1f}"
        
        running_mode_str = f"囗 连续(Continuously) &nbsp;&nbsp; 囗 走{run_min_str}分停{pause_min_str}分 &nbsp;&nbsp; 囗 正反转(forward+backward)"

        report_html = "<table border='1' cellspacing='0' cellpadding='6' style='border-collapse: collapse; width: 620px; font-size: 13px; text-align: center;'>"
        report_html += "<tr><td colspan='4' style='background-color: #DCDFE6; font-weight: bold; font-size: 15px;'>产品及基本信息</td></tr>"
        
        def add_tr(k1, v1, k2, v2):
            return f"<tr><td width='25%' style='background-color: #F0F2F5; font-weight: bold;'>{k1}</td><td width='25%'>{v1}</td><td width='25%' style='background-color: #F0F2F5; font-weight: bold;'>{k2}</td><td width='25%'>{v2}</td></tr>"

        dev_no = self.inputs_info["编号"].text() if "编号" in self.inputs_info else ""
        prod_name = self.inputs_info["产品名称"].text() if "产品名称" in self.inputs_info else ""
        prod_model = self.inputs_info["型号"].text() if "型号" in self.inputs_info else ""
        diam = self.inputs_info["轮径 (mm)"].text() if "轮径 (mm)" in self.inputs_info else ""
        mat = self.inputs_info["材质"].text() if "材质" in self.inputs_info else ""
        hard = self.inputs_info["硬度"].text() if "硬度" in self.inputs_info else ""
        remarks = self.inputs_info["备注"].text() if "备注" in self.inputs_info else ""

        report_html += add_tr("测试设备编号", dev_no, "Initiated By (提出者)", "")
        report_html += add_tr("Test Purpose (测试目的)", "", "Test Standard (测试标准)", "EN12532")
        report_html += add_tr("Description (产品名称)", prod_name, "Drawing (产品图号)", prod_model)
        report_html += add_tr("Diameter (外径)", f"{diam} mm" if diam else "", "Color (颜色)", "")
        report_html += add_tr("Material (材料)", mat, "Hardness (硬度)", hard)
        
        report_html += "<tr><td colspan='4' style='background-color: #DCDFE6; font-weight: bold; font-size: 15px;'>检验条件及数据快照</td></tr>"
        
        set_height_str = self.get_param_value_str("障碍高度 (mm)")
        display_height = f"{set_height_str} mm" if set_height_str and set_height_str != "0" else "无"

        report_html += add_tr("Running Time (已测时间)", snap['time'], "Trip (总行程)", snap['mileage'] + " km")
        report_html += add_tr("Speed (速度)", snap['speed'] + " km/h", "Test Load (检验负载)", f"1号:{snap['weight1']}kg / 2号:{snap['weight2']}kg")
        report_html += add_tr("障碍物高度", display_height, "碰撞障碍物次数/障碍次数", f"{display_actual_obs} / {self.get_param_value_str('障碍次数 (PCS)')} PCS")
        
        # 将 D38 的数值填入轮旋转总次数
        report_html += add_tr("NumberOfWheel Revolutions<br>(轮旋转的总次数)", str(snap['wheel_revolutions']), "Temperature (温度)", f"1号:{snap['temp1']}℃ / 2号:{snap['temp2']}℃")
        
        report_html += f"<tr><td style='background-color: #F0F2F5; font-weight: bold;'>Running mode (行走方式)</td><td colspan='3'>{running_mode_str}</td></tr>"
        
        report_html += "<tr><td colspan='4' style='background-color: #DCDFE6; font-weight: bold; font-size: 15px; text-align: left;'>Description Of Test Process(测试过程记录)</td></tr>"
        
        report_html += f"<tr><td colspan='2' rowspan='4'>&nbsp;</td><td colspan='2' rowspan='2'>备注：{remarks}</td></tr>"
        report_html += "<tr></tr>" 
        
        report_html += "<tr><td style='font-weight: bold;'>Conclusion(结论)：</td><td>囗 Accepted(合格） 囗 NO Accepted(不合格）</td></tr>" 
        report_html += f"<tr><td style='font-weight: bold;'>Date Of Test(日期):</td><td>{datetime.now().strftime('%Y.%m.%d')}</td></tr>" 
        
        report_html += "<tr><td style='font-weight: bold;'>Signature(审核)：</td><td>&nbsp;</td><td style='font-weight: bold;'>Operator(检验者):</td><td>&nbsp;</td></tr>" 

        report_html += "</table>"

        msg_box = QMessageBox(self)
        msg_box.setWindowTitle("动力检验报告 预览")
        msg_box.setTextFormat(Qt.RichText)
        msg_box.setText(report_html)
        msg_box.setIcon(QMessageBox.NoIcon)
        msg_box.exec_()

    # ==========================================
    # 核心：生成专属 Excel 报表
    # ==========================================
    def export_to_excel(self):
        if self.report_snapshot is None:
            self.capture_report_snapshot()
        snap = self.report_snapshot
        
        try:
            wb = Workbook()
            sheet = wb.active
            sheet.title = "动力检验报告"
            
            title_font = Font(name='宋体', size=20, bold=True)
            sub_title_font = Font(name='宋体', size=14, bold=True)
            header_font = Font(name='宋体', size=12, bold=True)
            normal_font = Font(name='宋体', size=11)
            
            center_align = Alignment(horizontal="center", vertical="center", wrapText=True)
            
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            for row in range(1, 22):
                for col in range(1, 5): 
                    cell = sheet.cell(row=row, column=col)
                    cell.border = thin_border
                    cell.alignment = center_align 
                    cell.font = normal_font

            sheet.merge_cells('A1:D1')
            sheet['A1'] = "中山市万里脚轮有限公司"
            sheet['A1'].font = title_font
            
            sheet.merge_cells('A2:D2')
            sheet['A2'] = "(脚轮及单轮 动力检验报告)"
            sheet['A2'].font = sub_title_font
            
            dev_no = self.inputs_info["编号"].text() if "编号" in self.inputs_info else ""
            prod_name = self.inputs_info["产品名称"].text() if "产品名称" in self.inputs_info else ""
            prod_model = self.inputs_info["型号"].text() if "型号" in self.inputs_info else ""
            diam = self.inputs_info["轮径 (mm)"].text() if "轮径 (mm)" in self.inputs_info else ""
            mat = self.inputs_info["材质"].text() if "材质" in self.inputs_info else ""
            hard = self.inputs_info["硬度"].text() if "硬度" in self.inputs_info else ""
            remarks = self.inputs_info["备注"].text() if "备注" in self.inputs_info else ""
            
            sheet['A3'] = "测试设备编号"
            sheet['B3'] = dev_no
            sheet['C3'] = "Initiated By:\n(提出者)"
            sheet['D3'] = ""
            
            sheet['A4'] = "Test Purpose:\n(测试目的)"
            sheet['B4'] = ""
            sheet['C4'] = "Test Standard:\n(测试标准)"
            sheet['D4'] = "EN12532"
            
            sheet['A5'] = "Description\n(产品名称)"
            sheet['B5'] = prod_name
            sheet['C5'] = "Drawing\n(产品图号)"
            sheet['D5'] = prod_model
            
            sheet['A6'] = "Diameter:\n(外径)"
            sheet['B6'] = f"{diam}mm" if diam else ""
            sheet['C6'] = "Color:\n(颜色)"
            sheet['D6'] = ""
            
            sheet['A7'] = "Material:\n(材料)"
            sheet['B7'] = mat
            sheet['C7'] = "Hardness:\n(硬度)"
            sheet['D7'] = hard
            
            sheet.merge_cells('A8:D8'); sheet['A8'] = "检测类型"; sheet['A8'].font = header_font
            
            sheet['A9'] = "囗 Type approval"
            sheet['B9'] = "囗 Compare To Competitor"
            sheet['C9'] = "囗 Special Test"
            sheet['D9'] = "囗 Production Test"
            
            sheet.merge_cells('A10:D10')
            sheet['A10'] = "Test conditions:(检验条件)"
            sheet['A10'].font = header_font
            
            set_obs_str = self.get_param_value_str("障碍数量 (PCS)")
            try:
                set_obs_val = int(set_obs_str) if set_obs_str.strip() else 0
            except ValueError:
                set_obs_val = 0
                
            # 核心逻辑：实测总次数 = 圈数 * 单圈障碍数
            actual_total_obstacles = snap['wheel_revolutions'] * set_obs_val
            display_actual_obs = f"{actual_total_obstacles} PCS" if actual_total_obstacles > 0 else "无"
            
            set_height_str = self.get_param_value_str("障碍高度 (mm)")
            display_height = f"{set_height_str} mm" if set_height_str and set_height_str != "0" else "无"
            
            th, tm, ts = self.time_inputs["测试时长"]
            run_s = th.value() * 3600 + tm.value() * 60 + ts.value()
            run_min = run_s / 60.0
            run_min_str = f"{int(run_min)}" if run_min.is_integer() else f"{run_min:.1f}"
            
            ih, im, i_s = self.time_inputs["间隔时间"]
            pause_s = ih.value() * 3600 + im.value() * 60 + i_s.value()
            pause_min = pause_s / 60.0
            pause_min_str = f"{int(pause_min)}" if pause_min.is_integer() else f"{pause_min:.1f}"
            
            sheet['A11'] = "Running Time:\n(已测时间)"
            sheet['B11'] = snap['time']
            sheet['C11'] = "Trip(km):\n(总行程)"
            sheet['D11'] = snap['mileage'] + " km"
            
            sheet['A12'] = "Speed:\n(速度)"
            sheet['B12'] = snap['speed'] + " km/h"
            sheet['C12'] = "Test Load:\n(检验负载)"
            sheet['D12'] = f"1号: {snap['weight1']}kg\n2号: {snap['weight2']}kg"
            
            cutoff_val = self.get_param_value_str("障碍次数 (PCS)")
            sheet['A13'] = "Height of obstacles:\n(障碍物高度)"
            sheet['B13'] = display_height
            sheet['C13'] = "Number of obstacles / Limit:\n(碰撞障碍物次数 / 障碍次数)"
            sheet['D13'] = f"{display_actual_obs} / {cutoff_val} PCS"
            
            # 【将 D38 获取的轮旋转次数填入】
            sheet['A14'] = "NumberOfWheel Revolutions\n(轮旋转的总次数)"
            sheet['B14'] = snap['wheel_revolutions']
            sheet['C14'] = "Temperature:\n(温度)"
            sheet['D14'] = f"1号: {snap['temp1']}℃\n2号: {snap['temp2']}℃"
            
            sheet['A15'] = "Running mode:\n(行走方式)"
            sheet.merge_cells('B15:D15') 
            sheet['B15'] = f"囗 连续(Continuously)      囗 走{run_min_str}分停{pause_min_str}分({run_min_str} min running + {pause_min_str} min pause)      囗 正反转(forward+backward)"
            
            sheet.merge_cells('A16:D16')
            sheet['A16'] = "Description Of Test Process(测试过程记录)"
            sheet['A16'].font = header_font
            sheet['A16'].alignment = Alignment(horizontal="center", vertical="center", wrapText=True)
            
            sheet.merge_cells('A17:B20')
            sheet['A17'] = ""
            
            sheet.merge_cells('C17:D18')
            sheet['C17'] = f"备注：{remarks}"
            
            sheet['C19'] = "Conclusion(结论)："
            sheet['D19'] = "囗 Accepted(合格） 囗 NO Accepted(不合格）"
            
            sheet['C20'] = "Date Of Test(日期):"
            now_date = datetime.now().strftime("%Y.%m.%d")
            sheet['D20'] = now_date
            
            sheet['A21'] = "Signature(审核)："
            sheet['B21'] = ""
            sheet['C21'] = "Operator(检验者):"
            sheet['D21'] = ""
            
            col_widths = {'A': 20, 'B': 40, 'C': 20, 'D': 45}
            for col, width in col_widths.items():
                sheet.column_dimensions[col].width = width

            sheet.row_dimensions[1].height = 40 
            sheet.row_dimensions[2].height = 30
            for row in range(3, 11):
                sheet.row_dimensions[row].height = 35
            for row in range(11, 15):
                sheet.row_dimensions[row].height = 60
            sheet.row_dimensions[15].height = 50 
            sheet.row_dimensions[16].height = 30 
            sheet.row_dimensions[17].height = 100
            sheet.row_dimensions[18].height = 40 
            for row in range(19, 22):
                sheet.row_dimensions[row].height = 35

            default_filename = f"万里脚轮测试报表_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path, _ = QFileDialog.getSaveFileName(self, "保存报表", default_filename, "Excel Files (*.xlsx)")

            if save_path:
                wb.save(save_path)
                QMessageBox.information(self, "生成成功", f"测试报表已成功生成并保存至：\n{save_path}")

        except PermissionError:
            QMessageBox.critical(self, "保存失败", "文件可能被其他程序占用，请关闭同名 Excel 文件后再试！")
        except Exception as e:
            QMessageBox.critical(self, "导出错误", f"生成 Excel 文件失败:\n{str(e)}")

    # ================= PLC 写入控制 =================
    def write_manual_register(self, address, value):
        def _write_task():
            try:
                temp_client = ModbusTcpClient(self.plc_ip, port=502, timeout=0.5)
                if temp_client.connect():
                    temp_client.write_register(address, value)
                    temp_client.close()
            except Exception as e:
                print(f"Async Write Failure D{address}: {e}")
        
        threading.Thread(target=_write_task, daemon=True).start()

    def import_params(self):
        try:
            val_mileage = float(self.inputs_params["测试里程 (km)"].text())
            val_speed = float(self.inputs_params["测试速度 (km/h)"].text())
            val_obs_cnt = int(self.inputs_params["障碍数量 (PCS)"].text())
            
            # 【注意：此处读取的就是界面的“障碍次数 (PCS)”，代表下发给 D52 的截至次数】
            val_cutoff = int(self.inputs_params["障碍次数 (PCS)"].text())
            
            val_obs_h = float(self.inputs_params["障碍高度 (mm)"].text())
            val_weight = float(self.inputs_params["承载重量 (kg)"].text())
            val_temp = float(self.inputs_params["承载温度 (℃)"].text())
            
        except ValueError:
            QMessageBox.warning(self, "格式错误", "参数格式不正确，请确保所有参数已正确输入数字！\n(不能包含字母或留空)")
            return

        try:
            builder = BinaryPayloadBuilder(byteorder=Endian.BIG, wordorder=Endian.LITTLE)
            builder.add_32bit_float(val_mileage) 
            builder.add_32bit_float(val_speed) 
            
            th, tm, ts = self.time_inputs["测试时长"]
            test_duration_s = th.value() * 3600 + tm.value() * 60 + ts.value()
            builder.add_32bit_int(test_duration_s)  
            
            ih, im, i_s = self.time_inputs["间隔时间"]
            interval_s = ih.value() * 3600 + im.value() * 60 + i_s.value()
            builder.add_32bit_int(interval_s)  
            
            # builder.add_16bit_int(val_obs_cnt) 
            # builder.add_16bit_int(0)
            builder.add_32bit_int(val_obs_cnt) # 改成32位int
            
            # builder.add_16bit_int(val_obs_h) 
            # builder.add_16bit_int(0)
            builder.add_32bit_float(val_obs_h) # 改成32位float
            
            builder.add_16bit_int(int(val_weight))
            builder.add_16bit_int(0) 
            
            builder.add_16bit_int(int(val_temp * 10)) 

            payload_main = builder.to_registers()
            
            builder_52 = BinaryPayloadBuilder(byteorder=Endian.BIG, wordorder=Endian.LITTLE)
            builder_52.add_32bit_int(val_cutoff)
            payload_52 = builder_52.to_registers()
            
            import_client = ModbusTcpClient(self.plc_ip, port=502, timeout=1)
            if import_client.connect():
                import_client.write_registers(18, payload_main)
                import_client.write_registers(52, payload_52)
                
                import_client.write_register(34, 1) 
                time.sleep(0.3)                     
                import_client.write_register(34, 0) 
                
                import_client.close()
                
                self.test_start_time = datetime.now().strftime("%H:%M:%S")
                self.report_snapshot = None 
                QMessageBox.information(self, "成功", "测试方案参数已成功导入 PLC！\n系统已记录测试开始时间。")
            else:
                QMessageBox.warning(self, "连接错误", "无法连接到 PLC。")
                
        except Exception as e:
            QMessageBox.critical(self, "通讯失败", f"无法写入PLC: {e}")

    def start_plc_thread(self):
        self.thread = PlcPollerThread(self.plc_ip)
        self.thread.data_updated.connect(self.update_plc_ui)
        self.thread.start()

    def closeEvent(self, event):
        self.save_settings()
        self.thread.stop()
        self.thread.wait(1500) 
        event.accept()

if __name__ == '__main__':
    if hasattr(Qt, 'AA_EnableHighDpiScaling'):
        QApplication.setAttribute(Qt.AA_EnableHighDpiScaling, True)
    if hasattr(Qt, 'AA_UseHighDpiPixmaps'):
        QApplication.setAttribute(Qt.AA_UseHighDpiPixmaps, True)

    app = QApplication(sys.argv)
    default_font = QApplication.font()
    default_font.setPointSize(10)
    app.setFont(default_font)
    
    window = MainWindow()
    window.show()
    sys.exit(app.exec_())